#!/usr/bin/env python3
# Por que: parse arquivos historicos BLS Table 1 RI (2000-2025) MANTENDO
# a hierarquia completa ate leaf-level. Emite CSV subitem-level pra
# reconstrucao bottom-up (Laspeyres Dez-anchor) analoga ao pareto_ipca.
# Complementa parse_historical_ri.py que so mantem os 41 agregados curados.
#
# Label -> code lookup: contra bls_maps/cu.item.tsv baixado de BLS FTP.
import re
import csv
import unicodedata
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "data" / "raw" / "relative_importance"
CU_ITEM = ROOT / "scripts" / "bls_maps" / "cu.item.tsv"
OUT = ROOT / "data" / "cpi_cpius_pesos_annual_subitem.csv"
OUT_UNMATCHED = ROOT / "data" / "cpi_cpius_pesos_annual_subitem_unmatched.csv"


def _norm(s):
    # Normaliza label pra lookup case-insensitive, punct-flex.
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower()
    # colapsa espacos, remove aspas
    s = re.sub(r"[\u2019']", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def load_cu_item():
    """Retorna dict {normalized_name: item_code}. Prioriza SA/SE sobre SS/AA."""
    rows = []
    with open(CU_ITEM, encoding="utf-8") as f:
        rdr = csv.DictReader(f, delimiter="\t")
        for r in rdr:
            code = r["item_code"].strip()
            name = r["item_name"].strip()
            level = int(r["display_level"].strip())
            rows.append((code, name, level))

    # Prefixo prioridade: SA > SE > SS > AA (AA e' old-base).
    prio = {"SA": 0, "SE": 1, "SS": 2, "AA": 9}
    rows.sort(key=lambda x: (prio.get(x[0][:2], 5), x[2]))

    lookup = {}
    for code, name, level in rows:
        key = _norm(name)
        if key not in lookup:
            lookup[key] = (code, name, level)
    return lookup


def _to_float(tok):
    tok = tok.strip()
    if not tok or tok == "-":
        return None
    try:
        return float(tok)
    except ValueError:
        return None


# ---------------- txt parser (2000-2019) ----------------
# Layout: `<spaces><label>...<dots>...<cpi_u>   <cpi_w>`.
# Indent = numero de espacos antes do primeiro char nao-espaco.
RE_TXT_ROW = re.compile(
    r"^(\s+)([A-Za-z][A-Za-z0-9,'()\-/ &.]+?)\s*\.{2,}\s*([\d.\-]+)\s+([\d.\-]+)\s*$"
)


def parse_txt(path):
    """
    Retorna lista de dicts {indent, label, cpi_u}.
    Trata label wrap: linha anterior soh com cauda do rotulo (sem dots).
    Trunca em 'Special aggregate indexes' (linha divisor).
    """
    out = []
    prev_label = None
    prev_indent = None
    with open(path, encoding="latin-1") as f:
        lines = f.readlines()

    for line in lines:
        line = line.rstrip("\n")
        # trunca no divisor de special aggregates
        if "Special aggregate indexes" in line:
            break
        m = RE_TXT_ROW.match(line)
        if m:
            indent_str, label, cu, _cw = m.groups()
            indent = len(indent_str)
            label = label.strip()
            # merge com wrap prev
            if prev_label and label[0].islower():
                label = prev_label + " " + label
                # indent do label real vem da linha wrap-tail (esta), nao da prev
            v = _to_float(cu)
            if v is not None:
                out.append({"indent": indent, "label": label, "cpi_u": v})
            prev_label = None
            prev_indent = None
        else:
            stripped = line.strip()
            # detecta label wrap: texto sem dots e sem valor
            if stripped and "..." not in line and not any(c.isdigit() for c in stripped):
                prev_label = stripped
                prev_indent = len(line) - len(line.lstrip())
    return out


# ---------------- xlsx parser (2020-2025) ----------------
def parse_xlsx(path):
    """
    Retorna lista de dicts {indent, label, cpi_u}.
    Trunca no divisor 'Special aggregate indexes' (label sem indent numerico).
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Table 1" not in wb.sheetnames:
        return []
    ws = wb["Table 1"]
    out = []
    for r in ws.iter_rows(min_row=1, values_only=True):
        if not r or len(r) < 3:
            continue
        indent, label, cu = r[0], r[1], r[2]
        if label and isinstance(label, str) and "Special aggregate" in label:
            break
        if not isinstance(indent, int):
            continue
        if not label or not isinstance(label, str):
            continue
        label = label.strip()
        if not isinstance(cu, (int, float)):
            continue
        out.append({"indent": indent, "label": label, "cpi_u": float(cu)})
    return out


# ---------------- label -> code lookup ----------------
# Fallbacks manuais pra drift historico bem conhecido.
# Chave = _norm(label do RI file), valor = _norm(label no cu.item).
LABEL_FALLBACKS = {
    # Renames em POF/hierarchy transitions ao longo dos anos.
    _norm("Owners' equivalent rent of primary residence"):
        _norm("Owners' equivalent rent of residences"),
    _norm("Utility (piped) gas"): _norm("Utility (piped) gas service"),
    _norm("Utility natural gas service"): _norm("Utility (piped) gas service"),
    _norm("Gas (piped) and electricity"): _norm("Energy services"),
    _norm("Airline fare"): _norm("Airline fares"),
    _norm("Hospital and related services"): _norm("Hospital services"),
    # Baby food -> Baby food and formula (2018+)
    _norm("Baby food"): _norm("Baby food and formula"),
    _norm("Sugar and artificial sweeteners"): _norm("Sugar and sugar substitutes"),
    _norm("Other pork including roasts and picnics"):
        _norm("Other pork including roasts, steaks, and ribs"),
    _norm("Other poultry including turkey"):
        _norm("Other uncooked poultry including turkey"),
    _norm("Men's furnishings"): _norm("Men's shirts and sweaters"),
    _norm("Women's underwear, nightwear, sportswear and accessories"):
        _norm("Women's underwear, nightwear, swimwear, and accessories"),
    _norm("Women's underwear, nightwear, swimwear, and accessories"):
        _norm("Women's underwear, nightwear, swimwear, and accessories"),
    _norm("Personal computers and peripheral equipment"):
        _norm("Computers, peripherals, and smart home assistant devices"),
    _norm("Photographers and film processing"):
        _norm("Photographers and photo processing"),
    _norm("Audio discs, tapes and other media"):
        _norm("Recorded music and music subscriptions"),
    _norm("Cable and satellite television and radio service"):
        _norm("Cable and satellite television service"),
    _norm("Prescription drugs and medical supplies"): _norm("Prescription drugs"),
    _norm("Nonprescription drugs and medical supplies"): _norm("Nonprescription drugs"),
    _norm("Nonprescription medical equipment and supplies"):
        _norm("Medical equipment and supplies"),
    _norm("Video discs and other media, including rental of video and audio"):
        _norm("Video discs and other media, including rental of video"),
    _norm("Video cassettes, discs, and other media including rental"):
        _norm("Video discs and other media, including rental of video"),
    _norm("Nursing homes and adult daycare"):
        _norm("Nursing homes and adult day services"),
    _norm("Club dues and fees for participant sports and group exercises"):
        _norm("Club membership for shopping clubs, fraternal, or other organizations, or participant sports fees"),
    _norm("Club membership dues and fees for participant sports"):
        _norm("Club membership for shopping clubs, fraternal, or other organizations, or participant sports fees"),
    _norm("Internal and respiratory over-the-counter drugs"): _norm("Nonprescription drugs"),
    _norm("Land-line telephone services, local charges"): _norm("Land-line telephone services"),
    _norm("Land-line telephone services, long distance charges"): _norm("Land-line telephone services"),
    _norm("State and local registration and license"):
        _norm("State motor vehicle registration and license fees"),
    # Sem correspondencia moderna ok — deixa unmatched:
    #   Housing at school (level 5), Technical and business school tuition (level 5),
    #   Care of invalids and elderly at home, Child care and nursery school,
    #   Fuels, Other household fuels, "Water and sewerage maintenance"
}


def resolve_label(label, lookup):
    key = _norm(label)
    if key in lookup:
        return lookup[key]
    if key in LABEL_FALLBACKS:
        alt = LABEL_FALLBACKS[key]
        if alt in lookup:
            return lookup[alt]
    # fuzzy: prefix match tolerando trailing terms
    # (evitar aqui — muitos false positives; melhor deixar unmatched e curar depois)
    return None


def collect_files():
    files = []
    for p in sorted((RAW / "ri-archive-2000-2009" / "2000-2009_RI_archive").glob("[0-9]*.txt")):
        if "old-weights" in p.name:
            continue
        files.append((int(p.stem), p, "txt"))
    for p in sorted((RAW / "ri-archive-2010-2019" / "2010-2019_RI_archive").glob("[0-9]*.txt")):
        if "old-weights" in p.name:
            continue
        files.append((int(p.stem), p, "txt"))
    for p in sorted(RAW.glob("[0-9]*.xlsx")):
        files.append((int(p.stem), p, "xlsx"))
    return files


def main():
    lookup = load_cu_item()
    print(f"[cu.item] {len(lookup)} distinct normalized names")

    files = collect_files()
    print(f"[parse] {len(files)} arquivos")

    matched_rows = []   # {year, indent, item_code, item_name, cpi_u}
    unmatched_rows = []  # {year, indent, label, cpi_u}

    for year, path, kind in files:
        parsed = parse_txt(path) if kind == "txt" else parse_xlsx(path)
        mch = 0
        for row in parsed:
            resolved = resolve_label(row["label"], lookup)
            if resolved is None:
                unmatched_rows.append({
                    "year": year, "indent": row["indent"],
                    "label": row["label"], "cpi_u": row["cpi_u"],
                })
                continue
            code, canonical, cu_level = resolved
            matched_rows.append({
                "year": year, "indent": row["indent"],
                "item_code": code, "item_name": canonical,
                "cpi_u": row["cpi_u"],
            })
            mch += 1
        print(f"  {year} ({kind}): {mch}/{len(parsed)} matched, {len(parsed)-mch} unmatched")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["year", "indent", "item_code", "item_name", "cpi_u"])
        w.writeheader()
        w.writerows(matched_rows)
    print(f"[out] {OUT}: {len(matched_rows)} linhas")

    if unmatched_rows:
        with open(OUT_UNMATCHED, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=["year", "indent", "label", "cpi_u"])
            w.writeheader()
            w.writerows(unmatched_rows)
        print(f"[unmatched] {OUT_UNMATCHED}: {len(unmatched_rows)} linhas")

        # sumario dos unmatched mais frequentes
        from collections import Counter
        c = Counter(r["label"] for r in unmatched_rows)
        print("\n[unmatched] top 20 rotulos:")
        for lbl, n in c.most_common(20):
            print(f"  {n:3d} x {lbl!r}")


if __name__ == "__main__":
    main()
